import numpy as np
import polars as pl
from collections import defaultdict, Counter
from datetime import date
from argparse import ArgumentParser as ap
from global_functions import date_2_index, index_2_date, city_to_id, get_list_of_weekends
import db_read as dbr
import config_loader as cfg
import yaml
from itertools import product
import os
import re
import matplotlib.pyplot as plt
from scipy.stats import binned_statistic_2d, linregress
from decimal import Decimal, InvalidOperation, localcontext, ROUND_HALF_EVEN
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import FixedLocator, LogLocator





db = dbr.db()

# ------------------- YAML laden -------------------#
def _load_yaml_data(filepaths=['tabelle_obs_for.yml']):
    config_data = {}
    for filepath in filepaths:
        with open(filepath, 'r', encoding='utf-8') as f: data = yaml.safe_load(f)
        if data:
            config_data.update(data)
    return config_data

# ------------------ gewünschte Genauigkeit -------------#

DEC_QUANT = Decimal('0.001') 

# ------------------- Hilfsfunktionen -------------------#
def to_value(raw_value):
    if raw_value is None:
        return None
    with localcontext() as ctx:
        ctx.prec = 12
        ctx.rounding = ROUND_HALF_EVEN
        d = Decimal(str(raw_value)) / Decimal('10')
        d = d.quantize(DEC_QUANT, rounding=ROUND_HALF_EVEN)
    return float(d)

def get_interval(value, ranges):
    for i, r in enumerate(ranges):
        a, b = r
        if a <= value <= b:
            return i, f"[{a}, {b}]"
    return None, None

intervals_cfg = _load_yaml_data(filepaths=['tabelle_obs_for.yml'])['Intervalle']
    
def get_obs_data(staedte, tage, elemente):
    obs_data = {}
    table_name = "wp_wetterturnier_obs"
    cursor = db.con.cursor(dictionary=True)
    for stadt in staedte:
        stationen = cfg.stationen[stadt]
        query = f"""
            SELECT station, betdate, p.paramName, value
            FROM {table_name} w
            INNER JOIN wp_wetterturnier_param p ON w.paramID = p.paramID
            WHERE station IN ({','.join(map(str, stationen))})
              AND betdate IN ({','.join(map(str, tage))})
              AND w.paramID IN ({','.join(map(str, elemente))})
            GROUP BY betdate, p.paramName, station
            ORDER BY betdate ASC, p.sort ASC, station ASC
        """
        cursor.execute(query)
        results = cursor.fetchall()
        nested = {}
        for row in results:
            betdate, param, raw_value = row['betdate'], row['paramName'], row['value']
            if raw_value is not None:
                nested.setdefault(betdate, {}).setdefault(param, []).append(to_value(raw_value))
        obs_data[cfg.id_zu_kuerzel[stadt]] = nested
    return obs_data   

                
def get_forecast_data(staedte, tage, elemente, users):
    forecast_data = {}
    table_name = "wp_wetterturnier_bets"
    cursor = db.con.cursor(dictionary=True)
    for stadt in staedte:
        query = f"""
            SELECT betdate, p.paramName, u.user_login, value
            FROM {table_name} w
            INNER JOIN wp_users u ON w.userID = u.ID
            INNER JOIN wp_wetterturnier_param p ON w.paramID = p.paramID
            WHERE w.cityID = {stadt}
              AND betdate IN ({','.join(map(str, tage))})
              AND w.paramID IN ({','.join(map(str, elemente))})
              AND w.userID IN ({','.join(map(str, users))})
            GROUP BY betdate, user_login, p.paramName
            ORDER BY betdate ASC, user_login ASC, p.sort ASC
        """
        cursor.execute(query)
        results = cursor.fetchall()
        nested = {}
        for row in results:
            betdate, user, param, raw_value = row['betdate'], row['user_login'], row['paramName'], row['value']
            if raw_value is not None:
                nested.setdefault(betdate, {}).setdefault(user, {})[param] = to_value(raw_value)
        forecast_data[cfg.id_zu_kuerzel[stadt]] = nested
    return forecast_data   


# ------------------- Klassenmittel -------------------#
def calc_class_means(intervals):
    means = []
    for r in intervals:
        with localcontext() as ctx:
            ctx.prec = 12
            ctx.rounding = ROUND_HALF_EVEN
            a, b = Decimal(str(r[0])), Decimal(str(r[1]))
            mean = (a + b) / 2
        means.append(mean)
    return means

# ------------------- Hauptprogramm -------------------
if __name__ == "__main__":
    db = dbr.db()
    ps = ap()
    ps.add_argument("--von", type=str, default=cfg.datum_neue_elemente)
    ps.add_argument("--bis", type=str, default=cfg.endtermin)
    ps.add_argument("-p", "--params", type=str, default=",".join(cfg.elemente_archiv_neu))
    ps.add_argument("-c", "--cities", type=str, default=",".join(cfg.stadtnamen))
    ps.add_argument("-u", "--users", type=str, default=",".join(cfg.auswertungsteilnehmer))
    ps.add_argument("-d", "--days", type=str, default=",".join(cfg.auswertungstage))
    ps.add_argument("--weighted", action="store_true", help="Calculate weighted BIAS")
    ps.add_argument("-v", "--verbose", action="store_true")
    ps = ps.parse_args()

    tdate_von = date_2_index(ps.von)
    tdate_bis = date_2_index(ps.bis)
    wochenendtage = get_list_of_weekends(tdate_von, tdate_bis)

    # Samstag oder Sonntag auswählen
    if ps.days == "Sat":
        selected_days = [d for d in wochenendtage if index_2_date(d).weekday() == 6]
    elif ps.days == "Sun":
        selected_days = [d for d in wochenendtage if index_2_date(d).weekday() == 0]
    else:
        selected_days = wochenendtage


    elemente_namen = [el for el in ps.params.split(",") if el in cfg.elemente_archiv_neu]
    elemente = db.get_param_ids(elemente_namen).values()
    staedte = [city_to_id(city, cfg) for city in ps.cities.split(",")]
    user_logins = ps.users.split(",")
    users_dict = db.get_user_ids(user_logins)
    users_dict_swapped = {v: k for k, v in users_dict.items()}
    users = users_dict.values()
    day_name = ps.days if ps.days in ["Sat","Sun"] else "All"

    # Daten laden – nur die ausgewählten Wochenendtage
    obs_data = get_obs_data(staedte, wochenendtage, elemente)
    forecast_data = get_forecast_data(staedte, wochenendtage, elemente, users)
    param_to_si_map = {name: unit for name, unit in zip(cfg.elemente_archiv_neu, cfg.elemente_einheiten_neu)}
    
# ------------------- Daten kombinieren -------------------#
    combined_data = {}
    for city in obs_data:
        combined_data[city] = {}
        for betdate in obs_data[city]:
            combined_data[city][betdate] = {
                'o': obs_data[city][betdate],
                'f': {}
            }
            for user in users:
                user_login = users_dict_swapped[user]
                user_login_actual = cfg.teilnehmerumbenennung.get(user_login, user_login)
                try:
                    combined_data[city][betdate]['f'][user_login] = forecast_data[city][betdate].get(
                        user_login_actual,
                        {el: None for el in elemente_namen}
                    )
                except KeyError:
                    combined_data[city][betdate]['f'][user_login] = {el: None for el in elemente_namen}

# ------------------- Verarbeitung und Export -------------------#


for param in elemente_namen:
    obs_ranges_def = intervals_cfg.get(param, [])
    for_ranges_def = intervals_cfg.get(param, [])
    if not obs_ranges_def or not for_ranges_def:
        print(f"Skipping {param} due to missing ranges.")
        continue

    obs_class_means = calc_class_means(obs_ranges_def)
    for_class_means = calc_class_means(for_ranges_def)

    counts = defaultdict(int)
    values_by_bin = defaultdict(list)
    for city, city_data in combined_data.items():
        for betdate, data in city_data.items():
            obs_vals_list = data['o'].get(param, [])
            valid_obs = [v for v in obs_vals_list if v is not None]
            if not valid_obs:
                continue
            obs_max = max(valid_obs)
            obs_idx, _ = get_interval(obs_max, obs_ranges_def)
            if obs_idx is None: continue
            obs_range_key = tuple(obs_ranges_def[obs_idx])

            for user, fvals in data['f'].items():
                fcast_val = fvals.get(param)
                if fcast_val is None: continue
                f_idx, _ = get_interval(fcast_val, for_ranges_def)
                if f_idx is None: continue
                for_range_key = tuple(for_ranges_def[f_idx])
                counts[(obs_range_key, for_range_key)] += 1
                values_by_bin[(obs_range_key, for_range_key)].append((obs_max, fcast_val))
    

# ------------------- Excel-Export komplett (Counts + Summen + Bias) ------------------- #


# Parameter aus YAML oder Datenstruktur
obs_classes = obs_ranges_def
fc_classes = for_ranges_def

n_rows = len(obs_classes)
n_cols = len(fc_classes)

# Ausgabe-Verzeichnis & Datei
all_city_str = "_".join(re.sub(r'[\\/:"*?<>|\s]+', '_', c) for c in combined_data.keys())
outdir = os.path.join("distribution_outputs", all_city_str)
os.makedirs(outdir, exist_ok=True)

users_set = {u for city_data in combined_data.values()
                 for betdate, data in city_data.items()
                 for u in data.get('f', {}).keys()}
user_str = "_".join(re.sub(r'[\\/:"*?<>|\s]+', '_', u) for u in users_set)

outfile_xlsx = os.path.join(outdir, f"distribution_all_{all_city_str}_{user_str}.xlsx")

# Workbook laden oder erstellen
if os.path.exists(outfile_xlsx):
    wb = load_workbook(outfile_xlsx)
else:
    wb = Workbook()
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
        wb.remove(wb["Sheet"])

# Neues Sheet
sheet_base_name = f"{param}_{all_city_str}"
sheet_name = sheet_base_name
counter = 1
while sheet_name in wb.sheetnames:
    sheet_name = f"{sheet_base_name}_{counter}"
    counter += 1
ws = wb.create_sheet(title=sheet_name)

# Styles
blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
orchid_fill = PatternFill(start_color="DA70D6", end_color="DA70D6", fill_type="solid")

# ------------------- Kopfzeilen ------------------- #
ws.cell(row=1, column=1, value="Kl")

# ------------------- Matrix Counts ------------------- #
matrix_counts = [[counts.get((tuple(obs_classes[i]), tuple(fc_classes[j])), 0)
                  for j in range(n_cols)] for i in range(n_rows)]

# Matrix in Excel schreiben
for i in range(n_rows):
    for j in range(n_cols):
        cell = ws.cell(row=i+2, column=j+2, value=matrix_counts[i][j])
        # Optional: Diagonale farbig markieren
        if i == j:
            cell.fill = blue_fill

# ------------------- Row- und Col-Summen ------------------- #
row_sums = [sum(row) for row in matrix_counts]
col_sums = [sum(matrix_counts[i][j] for i in range(n_rows)) for j in range(n_cols)]

for i, s in enumerate(row_sums):
    ws.cell(row=i+2, column=n_cols+2, value=s)
for j, s in enumerate(col_sums):
    ws.cell(row=n_rows+2, column=j+2, value=s)

ws.cell(row=n_rows+2, column=n_cols+2, value=sum(row_sums)).fill = orchid_fill



# MFc
for i in range(n_rows):
    fc_vals_all = []    # lege erstmal leere Listen für jede Zeile an
    for j in range(n_cols): # gehe nun durch jede Spalte (Vorhersageklassen)
        pairs = values_by_bin.get((tuple(obs_classes[i]), tuple(fc_classes[j])), [])
        fc_vals_all.extend([Decimal(str(f)) for o, f in pairs if f is not None])

    if fc_vals_all:
        fc_mfc = (sum(fc_vals_all) / Decimal(len(fc_vals_all))).quantize(Decimal('0.01'))   # gewichtetes Mittel
        ws.cell(row=i+2, column=1, value=fc_mfc)
    else:
        ws.cell(row=i+2, column=1, value='NIL')
# MOb
for j in range(n_cols):
    obs_vals_all = []
    for i in range(n_rows):
        pairs = values_by_bin.get((tuple(obs_classes[i]), tuple(fc_classes[j])), [])
        obs_vals_all.extend([Decimal(str(o)) for o, f in pairs if o is not None])

    if obs_vals_all:
        mob = (sum(obs_vals_all) / Decimal(len(obs_vals_all))).quantize(Decimal('0.01'))
        ws.cell(row=1, column=j+2, value=mob)
    else:
        ws.cell(row=1, column=j+2, value='NIL')


# ------------------- Summen-Beschriftungen ------------------- #
ws.cell(row=n_rows+2, column=1, value="Row_Sum")
ws.cell(row=1, column=n_cols+2, value="Col_Sum")
ws.cell(row=n_rows+3, column=1, value="BIAS")

# ------------------- Bias pro Forecast-Spalte ------------------- #
def calculate_bias(weighted=True):
    for j in range(n_cols):
        if col_sums[j] > 0:
            bias_sum = Decimal('0')
            for i in range(n_rows):
                mfci = ws.cell(row=i+2, column=1).value
                mobj = ws.cell(row=1, column=j+2).value

                if mfci in (None, 'NIL') or mobj in (None, 'NIL'):
                    continue
                try:
                    mfci_d = Decimal(str(mfci))
                    mobj_d = Decimal(str(mobj))
                    if weighted:
                        contrib = (mfci_d - mobj_d) * Decimal(matrix_counts[i][j]) / Decimal(col_sums[j])
                    else:
                        contrib = (mfci_d - mobj_d) / Decimal(n_rows)
                    bias_sum += contrib
                except (InvalidOperation, TypeError):
                    continue
            col_bias = bias_sum.quantize(Decimal('0.01'))
        else:
            col_bias = 'NIL'

        ws.cell(row=n_rows+3, column=j+2, value=float(col_bias) if col_bias != 'NIL' else 'NIL')

calculate_bias(weighted=ps.weighted if hasattr(ps, 'weighted') else False)

# ------------------- Excel speichern ------------------- #
wb.save(outfile_xlsx)
print(f"Excel table saved (sheet updated): {outfile_xlsx}")


outdir = os.path.join("distribution_outputs", all_city_str)

# ------------------- ASCII-Datei ----------------------- #

asc_outfile = os.path.join(outdir, f"distribution_{all_city_str}_{param}_{user_str}_{day_name}.asc")

col_widths_asc = [5, 6, 6, 4]
headers = ["Kl", "MFc", "MOb", "#"]  

asc_lines = [
    "  ".join(f"{h:>{w}}" for h, w in zip(headers, col_widths_asc)),
    "  ".join("-"*w for w in col_widths_asc)
]

    # --- jetzt Schleife über Forecast-Klassen ---
for j, (fc_lower, fc_upper) in enumerate(fc_classes):
    fc_vals = []
    obs_vals = []

        # alle Paare für diese Forecast-Klasse sammeln
    for i, obs_r in enumerate(obs_classes):
        pairs = values_by_bin.get((tuple(obs_r), tuple(fc_classes[j])), [])
        for o, f in pairs:
            if f is not None:
                fc_vals.append(Decimal(str(f)))
            if o is not None:
                obs_vals.append(Decimal(str(o)))

    # Mittelwerte
    mean_fc = (sum(fc_vals)/Decimal(len(fc_vals))) if fc_vals else "NIL"
    mean_obs = (sum(obs_vals)/Decimal(len(obs_vals))) if obs_vals else "NIL"

    # Counts
    obs_count = len(obs_vals)
    col_sum   = sum(matrix_counts[i][j] for i in range(n_cols))

    # Formatierung
    if mean_fc == "NIL":
        mf_format = f"{mean_fc:>{col_widths_asc[1]}}"
        mo_format = f"{mean_obs:>{col_widths_asc[2]}}"
    else:
        if fc_upper < 1.0:
            mf_format = f"{float(mean_fc):>{col_widths_asc[1]}.2f}"
            mo_format = f"{float(mean_obs):>{col_widths_asc[2]}.2f}"
        else:
            mf_format = f"{float(mean_fc):>{col_widths_asc[1]}.1f}"
            mo_format = f"{float(mean_obs):>{col_widths_asc[2]}.1f}"

    asc_lines.append("  ".join([
        f"{fc_upper:>{col_widths_asc[0]}.1f}",  # Forecast-Klassen-Maxima
        mf_format,
        mo_format,
        f"{col_sum:>{col_widths_asc[3]}}"       
    ]))

# Der ASCII-Teil wurde gänzlich korrigiert.



with open(asc_outfile, "w", encoding="utf-8") as f:
    f.write("\n".join(asc_lines))

print(f"Gesperichert unter: {asc_outfile}")


# ------------------- Plots ------------------- #
# Hier kommen die Plots. Hier habe ich die individuelle Skalierung für jeden Parameter eingefügt unter den vielen if's.
# Dann habe ich noch den dd12 Plot für jede Stadt als Polarkoordinatenplot hinzugefügt mit verschiedene Farben für die
# Obse und Forecasts.

plot_outdir = os.path.join(outdir, "plots")
os.makedirs(plot_outdir, exist_ok=True)

def set_linear_axis(ax, param):
    ax.set_xscale('linear')
    ax.set_yscale('linear')
    
    if param.lower() == "sd1":
        ticks = np.arange(0, 61, 10)
    elif param.lower() == "sd24":
        ticks = np.arange(0, 101, 20)
    elif param.lower() == "fx24":
        ticks = np.arange(5, 26, 5)
    else:
        ticks = None
    
    if ticks is not None:
        ax.xaxis.set_major_locator(FixedLocator(ticks))
        ax.yaxis.set_major_locator(FixedLocator(ticks))

# Scatterplots pro Parameter
for param in elemente_namen:
    obs_vals, fcast_vals = [], []

    # Daten sammeln: alle Werte nehmen, nicht nur Max
    for city, city_data in combined_data.items():
        for betdate, data in city_data.items():
            if betdate not in selected_days:
                continue
            obs_list = data["o"].get(param, [])
            if not obs_list:
                continue
            # nur das Maximum pro Beobachtung nehmen
            obs_max = max(obs_list)# max([v for v in obs_list if v is not None])
            for user, fvals in data["f"].items():
                fcast_val = fvals.get(param)
                if fcast_val is None:
                    continue
                obs_vals.append(obs_max)
                fcast_vals.append(fcast_val)


    if len(obs_vals) < 2:
        print(f"Not enough data for {param} to plot.")
        continue

    # Regression
    slope, intercept, r_value, _, _ = linregress(obs_vals, fcast_vals)

    # Frequenz pro Punkt berechnen (absolute Häufigkeit)
    counts = Counter(zip(obs_vals, fcast_vals))
    freqs = np.array([counts[(x, y)] for x, y in zip(obs_vals, fcast_vals)])

    # Scatterplot erstellen
    fig, ax = plt.subplots(figsize=(12, 8))
    set_linear_axis(ax, param)

    marker_size = 50  # Punktegröße

    # Achsenlimits bestimmen
    x_min, x_max = min(obs_vals), max(obs_vals)
    y_min, y_max = min(fcast_vals), max(fcast_vals)

    vmin = freqs.min()
    vmax = freqs.max()
    if vmin == vmax:
        vmin = 0
        vmax = freqs[0] + 1
        
        #Windrose
    if param == "dd12":
        obs_dirs_rad = np.deg2rad(obs_vals)
        fcast_dirs_rad = np.deg2rad(fcast_vals)

        # Plot
        fig = plt.figure(figsize=(8, 8))
        ax = fig.add_subplot(111, polar=True)
        n_bins = len(intervals_cfg.get(param, []))  # Anzahl Bins wie in YAML
        ax.hist(obs_dirs_rad, bins=n_bins, range=(0, 2*np.pi),
                alpha=0.6, color="blue", label="Obs")
        ax.hist(fcast_dirs_rad, bins=n_bins, range=(0, 2*np.pi),
                alpha=0.6, color="red", label="Forecast")

        ax.set_theta_zero_location("N")
        ax.set_theta_direction(-1)
        plt.legend()
        plt.title(f"wind direction distribution of {param} in {city}")

        plot_file_png = os.path.join(plot_outdir, f"windrose_{param}_{day_name}_{city}.png")
        plot_file_svg = os.path.join(plot_outdir, f"windrose_{param}_{day_name}_{city}.svg")
        plt.savefig(plot_file_png, dpi=300)
        plt.savefig(plot_file_svg)
        plt.close(fig)
        print(f"Windrichtungsplot gespeichert für {param}")
        print(f"Die Anzahl der bins ist: {n_bins}")
        continue


        
    else:    # Scatterplot
        scatter = ax.scatter(obs_vals, fcast_vals, c=freqs, s=marker_size,
                             cmap='coolwarm', alpha=0.7, vmin=vmin, vmax=vmax, clip_on=False)

        # Frequenz als Text anzeigen
        for x, y in zip(obs_vals, fcast_vals):
            freq = counts[(x, y)]
            ax.text(x, y, f"{freq}", fontsize=9, ha='center', va='center', color='black')

        # Colorbar erstellen
        cbar = plt.colorbar(scatter, ax=ax)
        cbar.set_label("Frequency (number of points)")
        if param.lower() in ["rr1", "rr24"]:
            ticks = np.arange(vmin, vmax+1, 10)
        else:
            ticks = np.arange(vmin, vmax+1, 1)
        cbar.set_ticks(ticks)
        cbar.set_ticklabels([f"{int(t)}" for t in ticks])

        # Achsenlimits & Linien
        if param.lower() == "sd1":
            ax.set_xlim(0, 60)
            ax.set_ylim(0, 60)
            ax.xaxis.set_major_locator(FixedLocator(np.arange(0, 61, 10)))
            ax.yaxis.set_major_locator(FixedLocator(np.arange(0, 61, 10)))
            cb_ticks = np.arange(0, 61, 10)  # Colorbar-Ticks
        elif param.lower() == "sd24":
            ax.set_xlim(0, 100)
            ax.set_ylim(0, 100)
            ax.xaxis.set_major_locator(FixedLocator(np.arange(0, 101, 20)))
            ax.yaxis.set_major_locator(FixedLocator(np.arange(0, 101, 20)))
            cb_ticks = np.arange(0, 101, 20)  # Colorbar-Ticks
        elif param.lower() == "ff12":
            ax.set_xlim(0, 15)
            ax.set_ylim(0, 15)
            ax.xaxis.set_major_locator(FixedLocator(np.arange(0, 16, 3)))
            ax.yaxis.set_major_locator(FixedLocator(np.arange(0, 16, 3)))
        elif param.lower() == "fx24":
            ax.set_xlim(0, 30)
            ax.set_ylim(0, 30)
            ax.xaxis.set_major_locator(FixedLocator(np.arange(0, 31, 5)))
            ax.yaxis.set_major_locator(FixedLocator(np.arange(0, 31, 5)))
        elif param.lower() == "tmin":
            ax.set_xlim(-15, 25)
            ax.set_ylim(-15, 25)
            ax.xaxis.set_major_locator(FixedLocator(np.arange(-15, 26, 5)))
            ax.yaxis.set_major_locator(FixedLocator(np.arange(-15, 26, 5)))
        elif param.lower() == "tmax":
            ax.set_xlim(-10, 40)
            ax.set_ylim(-10, 40)
            ax.xaxis.set_major_locator(FixedLocator(np.arange(-5, 41, 5)))
            ax.yaxis.set_major_locator(FixedLocator(np.arange(-5, 41, 5)))
        elif param.lower() == "td12":
            ax.set_xlim(-15, 25)
            ax.set_ylim(-15, 25)
            ax.xaxis.set_major_locator(FixedLocator(np.arange(-15, 26, 5)))
            ax.yaxis.set_major_locator(FixedLocator(np.arange(-15, 26, 5)))
        elif param.lower() == "rr1":
            ax.set_xscale("symlog", linthresh=0.1)  # linearer Bereich +-0.1
            ax.set_yscale("symlog", linthresh=0.1)
            ax.set_xlim(0, 100)
            ax.set_ylim(0, 100)

            # Major-Ticks: 1, 2, 5, 10, 20
            ax.xaxis.set_major_locator(LogLocator(base=10.0, subs=(1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0), numticks=10))
            ax.yaxis.set_major_locator(LogLocator(base=10.0, subs=(1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0), numticks=10))

        elif param.lower() == "rr24":
            ax.set_xscale("symlog", linthresh=0.1)
            ax.set_yscale("symlog", linthresh=0.1)
            ax.set_xlim(0, 200)
            ax.set_ylim(0, 200)

            # Major-Ticks: 1, 2, 5, 10, 20, 50, 100
            ax.xaxis.set_major_locator(LogLocator(base=10.0, subs=(1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0), numticks=10))
            ax.yaxis.set_major_locator(LogLocator(base=10.0, subs=(1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0), numticks=10))

        # Ursprungsgerrade Obs=Forecast
        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        ax.plot([xlim[0], xlim[1]], [xlim[0], xlim[1]], 'k--', label="Obs = Forecast")

        # Regressionslinie
        y_start = intercept + slope * xlim[0]
        y_end   = intercept + slope * xlim[1]
        ax.plot([xlim[0], xlim[1]], [y_start, y_end],
                'r-', label=rf"y = {slope:.2f}x + {intercept:.2f}, $R^2={r_value**2:.2f}$")


        # Achsenbeschriftung und Titel
        si_unit = param_to_si_map.get(param, "")
        ax.set_xlabel(f"Observation ({param}) [{si_unit}]")
        ax.set_ylabel(f"Forecast ({param}) [{si_unit}]")
        day_str = day_name if day_name in ["Sat", "Sun"] else "all days"
        ax.set_title(f"Scatterplot {param} for {day_str} and {', '.join(combined_data.keys())}")

        ax.grid(True)
        ax.legend()

        # Speichern
        
        
        plot_file_png = os.path.join(plot_outdir, f"scatter_{param}_{city}_{day_name}.png")
        plot_file_svg = os.path.join(plot_outdir, f"scatter_{param}_{city}_{day_name}.svg")
        plt.savefig(plot_file_png, dpi=300)
        plt.savefig(plot_file_svg)
        plt.show()
        plt.close(fig)

        print(f"Scatterplot saved for {param}")
        print("Total points:", len(obs_vals))
        print("Unique (obs, forecast) pairs:", len(counts))




















